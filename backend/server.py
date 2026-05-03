from dotenv import load_dotenv
from pathlib import Path
ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / ".env")

from fastapi import FastAPI, APIRouter, HTTPException, Request, Depends
from fastapi.responses import StreamingResponse, HTMLResponse
from motor.motor_asyncio import AsyncIOMotorClient

import os
import io
import csv
import json
import asyncio
import tempfile
import logging

from openpyxl import Workbook
from pydantic import BaseModel
from typing import List, Optional
from datetime import datetime, timezone, timedelta
from bson import ObjectId

import bcrypt
import jwt

# Google Drive
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload
from google_auth_oauthlib.flow import Flow
from google.oauth2.credentials import Credentials
from google.auth.transport.requests import Request as GoogleRequest


# ======================================================
# ENV / DB
# ======================================================

mongo_url = os.environ["MONGO_URL"]
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ["DB_NAME"]]

JWT_SECRET = os.environ["JWT_SECRET"]
JWT_ALGORITHM = "HS256"

from fastapi.middleware.cors import CORSMiddleware

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
    allow_credentials=True
)

api_router = APIRouter(prefix="/api")


origins = [
    "http://localhost:8081",
    "http://localhost:19006",
    "https://ahs-brown.vercel.app",
]


logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


# ======================================================
# HELPERS
# ======================================================

def serialize_doc(doc):
    if doc and "_id" in doc:
        doc["id"] = str(doc["_id"])
        del doc["_id"]
    return doc


def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode(), bcrypt.gensalt()).decode()


def verify_password(plain: str, hashed: str) -> bool:
    return bcrypt.checkpw(plain.encode(), hashed.encode())


def create_access_token(user_id: str, email: str):
    payload = {
        "sub": user_id,
        "email": email,
        "type": "access",
        "exp": datetime.now(timezone.utc) + timedelta(days=30)
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)


async def get_current_user(request: Request):
    auth = request.headers.get("Authorization", "")

    if not auth.startswith("Bearer "):
        raise HTTPException(status_code=401, detail="Unauthorized")

    token = auth[7:]

    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        user = await db.users.find_one({"_id": ObjectId(payload["sub"])})
        if not user:
            raise HTTPException(status_code=401, detail="User not found")

        return {
            "id": str(user["_id"]),
            "email": user["email"],
            "name": user.get("name", "")
        }

    except:
        raise HTTPException(status_code=401, detail="Invalid token")


# ======================================================
# MODELS
# ======================================================

class LoginRequest(BaseModel):
    email: str
    password: str


class TransactionCreate(BaseModel):
    date: str
    amount: float
    type: str
    mode: str
    category: Optional[str] = None
    description: Optional[str] = ""


# ======================================================
# SETTINGS
# ======================================================

async def get_settings():
    s = await db.settings.find_one()

    if not s:
        await db.settings.insert_one({
            "bank_balance": 0,
            "petty_cash_balance": 0
        })
        s = await db.settings.find_one()

    return s


async def update_balance(field: str, amount: float):
    await db.settings.update_one({}, {"$inc": {field: amount}}, upsert=True)

async def recalculate_balances():
    txns = await db.transactions.find().to_list(10000)

    bank = 0
    petty = 0

    for t in txns:
        amt = float(t.get("amount", 0))
        sign = 1 if t.get("type") == "Income" else -1

        if t.get("mode") == "Bank":
            bank += amt * sign

        elif t.get("mode") == "Petty Cash":
            petty += amt * sign

    await db.settings.update_one(
        {},
        {
            "$set": {
                "bank_balance": bank,
                "petty_cash_balance": petty
            }
        },
        upsert=True
    )

# ======================================================
# AUTH
# ======================================================

@api_router.post("/auth/login")
async def login(req: LoginRequest):
    user = await db.users.find_one({"email": req.email.lower().strip()})

    if not user:
        raise HTTPException(status_code=401, detail="Invalid login")

    if not verify_password(req.password, user["password_hash"]):
        raise HTTPException(status_code=401, detail="Invalid login")

    token = create_access_token(str(user["_id"]), user["email"])

    return {
        "token": token,
        "user": {
            "id": str(user["_id"]),
            "email": user["email"],
            "name": user.get("name", "")
        }
    }


@api_router.get("/auth/me")
async def auth_me(user=Depends(get_current_user)):
    return user


# ======================================================
# DASHBOARD
# ======================================================

@api_router.get("/dashboard")
async def dashboard(user=Depends(get_current_user)):
    settings = await get_settings()

    txns = await db.transactions.find().to_list(10000)
    projects = await db.projects.find().to_list(10000)
    partners = await db.partners.find().to_list(10000)
    inventory = await db.inventory.find().to_list(10000)

    income = sum(float(t.get("amount", 0)) for t in txns if t.get("type") == "Income")
    expense = sum(float(t.get("amount", 0)) for t in txns if t.get("type") == "Expense")

    receivables = 0
    for p in projects:
        total = float(p.get("total_amount", 0) or 0)
        received = float(p.get("received_amount", 0) or 0)
        receivables += max(total - received, 0)

    partner_balance = sum(float(p.get("balance", 0) or 0) for p in partners)
    stock_total = sum(float(i.get("stock", 0) or 0) for i in inventory)

    recent_bank = [
        serialize_doc(t) for t in txns
        if t.get("mode") == "Bank"
    ][-5:]

    return {
        "bank_balance": settings.get("bank_balance", 0),
        "petty_cash_balance": settings.get("petty_cash_balance", 0),
        "total_balance":
            settings.get("bank_balance", 0) +
            settings.get("petty_cash_balance", 0),

        "total_income": income,
        "total_expenses": expense,
        "profit_loss": income - expense,

        "receivables": receivables,
        "partner_balance_total": partner_balance,
        "inventory_stock_total": stock_total,
        "recent_bank_transactions": recent_bank,

        "drive_connected":
            await db.drive_credentials.find_one({"user_id": user["id"]})
            is not None
    }


# ======================================================
# TRANSACTIONS
# ======================================================

@api_router.get("/transactions")
async def transactions(user=Depends(get_current_user)):
    txns = await db.transactions.find().sort("date", -1).to_list(10000)
    return [serialize_doc(t) for t in txns]


@api_router.post("/transactions")
async def add_transaction(
    data: TransactionCreate,
    user=Depends(get_current_user)
):
    d = data.dict()
    d["created_at"] = datetime.now(timezone.utc)

    result = await db.transactions.insert_one(d)

    if d["mode"] == "Bank":
        await update_balance(
            "bank_balance",
            d["amount"] if d["type"] == "Income" else -d["amount"]
        )

    elif d["mode"] == "Petty Cash":
        await update_balance(
            "petty_cash_balance",
            d["amount"] if d["type"] == "Income" else -d["amount"]
        )

    created = await db.transactions.find_one({"_id": result.inserted_id})
    return serialize_doc(created)

@api_router.put("/transactions/{item_id}")
async def update_transaction(item_id: str, data: dict, user=Depends(get_current_user)):
    await db.transactions.update_one(
        {"_id": ObjectId(item_id)},
        {"$set": data}
    )

    await recalculate_balances()

    row = await db.transactions.find_one({"_id": ObjectId(item_id)})
    return serialize_doc(row)


@api_router.delete("/transactions/{item_id}")
async def delete_transaction(item_id: str, user=Depends(get_current_user)):
    await db.transactions.delete_one({"_id": ObjectId(item_id)})

    await recalculate_balances()

    return {"message": "Deleted"}

# ======================================================
# EXPORT CSV
# ======================================================

@api_router.get("/export/transactions")
async def export_transactions(user=Depends(get_current_user)):
    txns = await db.transactions.find().sort("date", -1).to_list(10000)

    output = io.StringIO()
    writer = csv.writer(output)

    writer.writerow(
        ["Date", "Type", "Mode", "Category", "Amount", "Description"]
    )

    for t in txns:
        writer.writerow([
            t.get("date", ""),
            t.get("type", ""),
            t.get("mode", ""),
            t.get("category", ""),
            t.get("amount", 0),
            t.get("description", "")
        ])

    output.seek(0)

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={
            "Content-Disposition":
                "attachment; filename=transactions.csv"
        }
    )


# ======================================================
# GOOGLE DRIVE
# ======================================================

def get_drive_flow():
    redirect_uri = os.environ["GOOGLE_DRIVE_REDIRECT_URI"]

    client_config = {
        "web": {
            "client_id": os.environ["GOOGLE_CLIENT_ID"],
            "client_secret": os.environ["GOOGLE_CLIENT_SECRET"],
            "auth_uri": "https://accounts.google.com/o/oauth2/auth",
            "token_uri": "https://oauth2.googleapis.com/token",
            "redirect_uris": [redirect_uri]
        }
    }

    return Flow.from_client_config(
        client_config,
        scopes=["https://www.googleapis.com/auth/drive.file"],
        redirect_uri=redirect_uri
    )


async def get_drive_service_for_user(user_id: str):
    creds_doc = await db.drive_credentials.find_one({"user_id": user_id})

    if not creds_doc:
        return None

    creds = Credentials(
        token=creds_doc["access_token"],
        refresh_token=creds_doc.get("refresh_token"),
        token_uri=creds_doc["token_uri"],
        client_id=creds_doc["client_id"],
        client_secret=creds_doc["client_secret"],
        scopes=creds_doc["scopes"]
    )

    if creds.expired and creds.refresh_token:
        creds.refresh(GoogleRequest())

        await db.drive_credentials.update_one(
            {"user_id": user_id},
            {"$set": {"access_token": creds.token}}
        )

    return build("drive", "v3", credentials=creds)


@api_router.get("/drive/connect")
async def drive_connect(user=Depends(get_current_user)):
    flow = get_drive_flow()

    auth_url, state = flow.authorization_url(
        access_type="offline",
        prompt="consent",
        include_granted_scopes="true"
    )

    await db.oauth_temp.delete_many({"user_id": user["id"]})

    await db.oauth_temp.insert_one({
        "user_id": user["id"],
        "state": state,
        "code_verifier": flow.code_verifier
    })

    return {"authorization_url": auth_url}


@api_router.get("/oauth/drive/callback")
async def drive_callback(code: str, state: str = ""):
    try:
        saved = await db.oauth_temp.find_one({"state": state})

        if not saved:
            return HTMLResponse("<h2>Connection Failed</h2>")

        flow = get_drive_flow()
        flow.code_verifier = saved["code_verifier"]

        flow.fetch_token(code=code)

        creds = flow.credentials

        await db.drive_credentials.update_one(
            {"user_id": saved["user_id"]},
            {
                "$set": {
                    "user_id": saved["user_id"],
                    "access_token": creds.token,
                    "refresh_token": creds.refresh_token,
                    "token_uri": creds.token_uri,
                    "client_id": creds.client_id,
                    "client_secret": creds.client_secret,
                    "scopes": list(creds.scopes)
                }
            },
            upsert=True
        )

        await db.oauth_temp.delete_many({"state": state})

        return HTMLResponse("""
        <html>
        <body>
        <h2>Google Drive Connected</h2>
        <script>window.close();</script>
        </body>
        </html>
        """)

    except Exception as e:
        return HTMLResponse(f"<h2>Connection Failed</h2><p>{str(e)}</p>")


# ======================================================
# BACKUP ONLY (.xlsx)
# ======================================================

async def run_drive_backup(user_id: str):
    service = await get_drive_service_for_user(user_id)

    if not service:
        raise HTTPException(status_code=400, detail="Drive not connected")

    folder_name = "Aruvi Housing Solutions - Backup"

    query = (
        f"name='{folder_name}' and "
        "mimeType='application/vnd.google-apps.folder' "
        "and trashed=false"
    )

    folders = service.files().list(
        q=query,
        spaces="drive",
        fields="files(id,name)"
    ).execute().get("files", [])

    if folders:
        folder_id = folders[0]["id"]
    else:
        created = service.files().create(
            body={
                "name": folder_name,
                "mimeType": "application/vnd.google-apps.folder"
            },
            fields="id"
        ).execute()

        folder_id = created["id"]

    wb = Workbook()

    # Transactions
    ws = wb.active
    ws.title = "Transactions"
    ws.append(["Date", "Type", "Mode", "Category", "Amount", "Description"])

    txns = await db.transactions.find().to_list(10000)
    for t in txns:
        ws.append([
            t.get("date", ""),
            t.get("type", ""),
            t.get("mode", ""),
            t.get("category", ""),
            t.get("amount", 0),
            t.get("description", "")
        ])

    # Projects
    ws2 = wb.create_sheet("Projects")
    ws2.append(["Name", "Total Amount", "Received Amount"])

    projects = await db.projects.find().to_list(10000)
    for p in projects:
        ws2.append([
            p.get("name", ""),
            p.get("total_amount", 0),
            p.get("received_amount", 0)
        ])

    # Partners
    ws3 = wb.create_sheet("Partners")
    ws3.append(["Name", "Balance"])

    partners = await db.partners.find().to_list(10000)
    for p in partners:
        ws3.append([
            p.get("name", ""),
            p.get("balance", 0)
        ])

    # Inventory
    ws4 = wb.create_sheet("Inventory")
    ws4.append(["Bag Type", "Stock"])

    inv = await db.inventory.find().to_list(10000)
    for r in inv:
        ws4.append([
            r.get("bag_type", ""),
            r.get("stock", 0)
        ])

    # Inventory Purchases
    ws5 = wb.create_sheet("Inventory Purchases")
    ws5.append(["Date", "Bag Type", "Quantity"])

    pur = await db.inventory_purchases.find().to_list(10000)
    for r in pur:
        ws5.append([
            r.get("date", ""),
            r.get("bag_type", ""),
            r.get("quantity", 0)
        ])

    filename = datetime.now().strftime(
        "AHS_Backup_%d-%m-%Y_%H-%M.xlsx"
    )

    with tempfile.NamedTemporaryFile(
        suffix=".xlsx",
        delete=False
    ) as tmp:

        wb.save(tmp.name)

        media = MediaFileUpload(
            tmp.name,
            mimetype=(
                "application/vnd.openxmlformats-"
                "officedocument.spreadsheetml.sheet"
            )
        )

        service.files().create(
            body={
                "name": filename,
                "parents": [folder_id]
            },
            media_body=media
        ).execute()

    await db.backup_log.insert_one({
        "user_id": user_id,
        "timestamp": datetime.now(timezone.utc),
        "file": filename
    })

    return {
        "message": "Backup completed",
        "file": filename
    }


@api_router.post("/drive/backup")
async def backup(user=Depends(get_current_user)):
    return await run_drive_backup(user["id"])


@api_router.get("/drive/disconnect")
async def disconnect(user=Depends(get_current_user)):
    await db.drive_credentials.delete_many({"user_id": user["id"]})
    return {"message": "Disconnected"}

# =========================
# PROJECTS
# =========================

@api_router.get("/projects")
async def get_projects(user=Depends(get_current_user)):
    rows = await db.projects.find().sort("_id", -1).to_list(1000)
    return [serialize_doc(x) for x in rows]

@api_router.post("/projects")
async def create_project(data: dict, user=Depends(get_current_user)):
    data["created_at"] = datetime.now(timezone.utc)
    data.setdefault("received_amount", 0)
    data.setdefault("bag_usage_history", [])
    r = await db.projects.insert_one(data)
    row = await db.projects.find_one({"_id": r.inserted_id})
    return serialize_doc(row)

@api_router.put("/projects/{item_id}")
async def update_project(item_id: str, data: dict, user=Depends(get_current_user)):
    await db.projects.update_one(
        {"_id": ObjectId(item_id)},
        {"$set": data}
    )
    row = await db.projects.find_one({"_id": ObjectId(item_id)})
    return serialize_doc(row)

@api_router.delete("/projects/{item_id}")
async def delete_project(item_id: str, user=Depends(get_current_user)):
    await db.projects.delete_one({"_id": ObjectId(item_id)})
    return {"message": "Deleted"}

@api_router.get("/projects/{item_id}")
async def get_project(item_id: str, user=Depends(get_current_user)):
    row = await db.projects.find_one({"_id": ObjectId(item_id)})
    if not row:
        raise HTTPException(status_code=404, detail="Not found")
    return serialize_doc(row)

@api_router.post("/projects/{item_id}/bag-usage")
async def add_bag_usage(item_id: str, data: dict, user=Depends(get_current_user)):
    qty = int(data.get("quantity", 0))
    bag_type = data.get("bag_type", "Naturoplast")

    await db.projects.update_one(
        {"_id": ObjectId(item_id)},
        {"$push": {
            "bag_usage_history": {
                "date": data.get("date"),
                "bag_type": bag_type,
                "quantity": qty
            }
        }}
    )

    await db.inventory.update_one(
        {"bag_type": bag_type},
        {"$inc": {"stock": -qty}},
        upsert=True
    )

    row = await db.projects.find_one({"_id": ObjectId(item_id)})
    return serialize_doc(row)

# =========================
# PARTNERS
# =========================

@api_router.get("/partners")
async def get_partners(user=Depends(get_current_user)):
    rows = await db.partners.find().sort("_id", -1).to_list(1000)
    result = []
    for p in rows:
        p = serialize_doc(p)
        # Compute totals from partner_transactions collection
        pid = p["id"]
        all_txns = await db.partner_transactions.find({"partner_id": pid}).to_list(10000)
        total_invested = sum(
            float(t.get("amount", 0)) for t in all_txns if t.get("type") == "Investment"
        )
        total_withdrawn = sum(
            float(t.get("amount", 0)) for t in all_txns if t.get("type") == "Withdrawal"
        )
        current_balance = total_invested - total_withdrawn
        p["total_investment"] = total_invested
        p["total_withdrawals"] = total_withdrawn
        p["current_balance"] = current_balance
        p["balance"] = current_balance
        result.append(p)
    return result

@api_router.post("/partners")
async def create_partner(data: dict, user=Depends(get_current_user)):
    data.setdefault("balance", 0)
    r = await db.partners.insert_one(data)
    row = await db.partners.find_one({"_id": r.inserted_id})
    p = serialize_doc(row)
    p["total_investment"] = float(data.get("total_investment", 0))
    p["total_withdrawals"] = 0.0
    p["current_balance"] = p["total_investment"]
    return p

@api_router.put("/partners/{item_id}")
async def update_partner(item_id: str, data: dict, user=Depends(get_current_user)):
    await db.partners.update_one(
        {"_id": ObjectId(item_id)},
        {"$set": data}
    )
    row = await db.partners.find_one({"_id": ObjectId(item_id)})
    return serialize_doc(row)

@api_router.delete("/partners/{item_id}")
async def delete_partner(item_id: str, user=Depends(get_current_user)):
    await db.partners.delete_one({"_id": ObjectId(item_id)})
    return {"message": "Deleted"}


# FIX: Route now matches what the frontend calls: POST /api/partners/{id}/transaction
@api_router.post("/partners/{partner_id}/transaction")
async def partner_txn(partner_id: str, data: dict, user=Depends(get_current_user)):
    amt = float(data.get("amount", 0))

    # Accept both "Investment"/"Withdrawal" and "invest"/"withdraw" from frontend
    raw_type = data.get("type", "Investment")
    if raw_type in ("invest", "Investment"):
        txn_type = "Investment"
    else:
        txn_type = "Withdrawal"

    date = data.get("date", datetime.now(timezone.utc).strftime("%Y-%m-%d"))

    # Update partner balance
    delta = amt if txn_type == "Investment" else -amt
    await db.partners.update_one(
        {"_id": ObjectId(partner_id)},
        {"$inc": {"balance": delta}}
    )

    # Save in partner_transactions with partner_id as string
    await db.partner_transactions.insert_one({
        "partner_id": partner_id,
        "amount": amt,
        "type": txn_type,
        "date": date,
        "created_at": datetime.now(timezone.utc)
    })

    # Save in main transactions
    txn = {
        "date": date,
        "amount": amt,
        "type": "Income" if txn_type == "Investment" else "Expense",
        "mode": "Bank",
        "category": "Partner",
        "description": f"{txn_type} - Partner",
        "created_at": datetime.now(timezone.utc)
    }
    await db.transactions.insert_one(txn)

    # Update bank balance
    await update_balance(
        "bank_balance",
        amt if txn["type"] == "Income" else -amt
    )

    return {"message": "Saved", "type": txn_type, "amount": amt}


# Keep old route for backward compatibility (no-op redirect)
@api_router.post("/partners/transaction")
async def partner_txn_legacy(data: dict, user=Depends(get_current_user)):
    pid = data.get("partner_id")
    if not pid:
        raise HTTPException(status_code=400, detail="partner_id required")
    return await partner_txn(pid, data, user)


# =========================
# INVENTORY
# =========================

@api_router.get("/inventory")
async def get_inventory(user=Depends(get_current_user)):
    # FIX: Return a computed summary object instead of raw array
    inv_rows = await db.inventory.find().to_list(1000)
    purchases = await db.inventory_purchases.find().to_list(10000)

    # Compute per-type stock
    naturoplast_stock = 0
    iraniya_stock = 0
    for row in inv_rows:
        bt = row.get("bag_type", "")
        stock = int(row.get("stock", 0))
        if bt == "Naturoplast":
            naturoplast_stock = stock
        elif bt == "Iraniya":
            iraniya_stock = stock

    # Compute purchased totals from purchase history
    naturoplast_purchased = sum(
        int(p.get("bags", p.get("quantity", 0)))
        for p in purchases if p.get("bag_type") == "Naturoplast"
    )
    iraniya_purchased = sum(
        int(p.get("bags", p.get("quantity", 0)))
        for p in purchases if p.get("bag_type") == "Iraniya"
    )
    total_purchased = naturoplast_purchased + iraniya_purchased

    # Compute used from all project bag_usage_history
    projects = await db.projects.find().to_list(10000)
    naturoplast_used = 0
    iraniya_used = 0
    for proj in projects:
        for usage in proj.get("bag_usage_history", []):
            # Guard against corrupted string entries in bag_usage_history
            if not isinstance(usage, dict):
                continue
            qty = int(usage.get("quantity", 0))
            if usage.get("bag_type") == "Naturoplast":
                naturoplast_used += qty
            elif usage.get("bag_type") == "Iraniya":
                iraniya_used += qty
    total_used = naturoplast_used + iraniya_used

    current_stock = naturoplast_stock + iraniya_stock

    # Build purchase history list for display
    purchase_history = []
    for p in sorted(purchases, key=lambda x: x.get("date", ""), reverse=True):
        purchase_history.append({
            "id": str(p["_id"]),          # <-- THIS LINE IS THE KEY ADDITION
            "date": p.get("date", ""),
            "bag_type": p.get("bag_type", ""),
            "bags": int(p.get("bags", p.get("quantity", 0))),
            "amount": float(p.get("amount", 0)),
            "mode": p.get("mode", "Bank"),
        })

    return {
        "current_stock": current_stock,
        "naturoplast_stock": naturoplast_stock,
        "iraniya_stock": iraniya_stock,
        "naturoplast_purchased": naturoplast_purchased,
        "iraniya_purchased": iraniya_purchased,
        "naturoplast_used": naturoplast_used,
        "iraniya_used": iraniya_used,
        "total_purchased": total_purchased,
        "total_used": total_used,
        "purchase_history": purchase_history,
    }

@api_router.post("/inventory/purchase")
async def add_inventory_purchase(data: dict, user=Depends(get_current_user)):
    # Accept both "bags" and "quantity" field names
    qty = int(data.get("bags", data.get("quantity", 0)))
    bag_type = data.get("bag_type", "Naturoplast")
    amount = float(data.get("amount", 0))
    mode = data.get("mode", "Bank")
    date = data.get("date", datetime.now(timezone.utc).strftime("%Y-%m-%d"))

    # Update inventory stock
    await db.inventory.update_one(
        {"bag_type": bag_type},
        {"$inc": {"stock": qty}},
        upsert=True
    )

    # Save purchase record (normalize to use "bags" field)
    await db.inventory_purchases.insert_one({
        "bags": qty,
        "quantity": qty,
        "bag_type": bag_type,
        "amount": amount,
        "mode": mode,
        "date": date,
        "created_at": datetime.now(timezone.utc)
    })

    # Also record as an expense transaction
    txn = {
        "date": date,
        "amount": amount,
        "type": "Expense",
        "mode": mode,
        "category": "Inventory",
        "description": f"Bag Purchase - {bag_type} ({qty} bags)",
        "created_at": datetime.now(timezone.utc)
    }
    await db.transactions.insert_one(txn)

    if mode == "Bank":
        await update_balance("bank_balance", -amount)
    elif mode == "Petty Cash":
        await update_balance("petty_cash_balance", -amount)

    return {"message": "Saved"}

@api_router.get("/inventory-purchases")
async def get_inventory_purchases(user=Depends(get_current_user)):
    rows = await db.inventory_purchases.find().sort("_id", -1).to_list(1000)
    return [serialize_doc(x) for x in rows]

# =======================================================
# ADD THIS BLOCK to server.py
# Paste it right after the existing @api_router.post("/inventory/purchase") block
# =======================================================

@api_router.delete("/inventory/purchase/{purchase_id}")
async def delete_inventory_purchase(purchase_id: str, user=Depends(get_current_user)):
    # Find the purchase record first
    purchase = await db.inventory_purchases.find_one({"_id": ObjectId(purchase_id)})
    if not purchase:
        raise HTTPException(status_code=404, detail="Purchase not found")

    qty = int(purchase.get("bags", purchase.get("quantity", 0)))
    bag_type = purchase.get("bag_type", "Naturoplast")
    amount = float(purchase.get("amount", 0))
    mode = purchase.get("mode", "Bank")

    # Deduct from inventory stock
    await db.inventory.update_one(
        {"bag_type": bag_type},
        {"$inc": {"stock": -qty}}
    )

    # Delete the purchase record
    await db.inventory_purchases.delete_one({"_id": ObjectId(purchase_id)})

    # Also reverse the expense transaction if it exists
    # (find by description match and amount - best effort)
    await db.transactions.delete_one({
        "category": "Inventory",
        "amount": amount,
        "mode": mode,
        "description": f"Bag Purchase - {bag_type} ({qty} bags)"
    })

    # Recalculate bank/petty cash balances
    await recalculate_balances()

    return {"message": "Purchase deleted and stock updated"}


# ======================================================
# STARTUP
# ======================================================

@app.on_event("startup")
async def startup():
    email = os.environ.get("ADMIN_EMAIL", "admin@example.com").lower()
    password = os.environ.get("ADMIN_PASSWORD", "admin123")

    existing = await db.users.find_one({"email": email})

    if not existing:
        await db.users.insert_one({
            "email": email,
            "password_hash": hash_password(password),
            "name": "Aruvi Housing Solutions",
            "role": "admin"
        })

    await db.users.create_index("email", unique=True)

    asyncio.create_task(auto_backup_scheduler())


async def auto_backup_scheduler():
    while True:
        await asyncio.sleep(86400)

        try:
            admin = await db.users.find_one({"role": "admin"})

            if admin:
                creds = await db.drive_credentials.find_one({
                    "user_id": str(admin["_id"])
                })

                if creds:
                    await run_drive_backup(str(admin["_id"]))

        except Exception as e:
            logger.error(e)


# ======================================================
# APP
# ======================================================

app.include_router(api_router)


@app.on_event("shutdown")
async def shutdown():
    client.close()
